# -*- coding:utf-8 -*-
"""
-------------------------------------------
@Time : 2020/5/9 12:09 
@Auth : 杨哲
@File : python_handle_excel.py
@IDE : PyCharm
@Motto : Never Stop Learning
-------------------------------------------
"""
from openpyxl import load_workbook
from scripts.handle_config import testcase_config
from scripts.constants import DATA_DIR
import os


class HandleExcel:
    '''
    处理文件类
    '''

    def __init__(self, filename, sheetname=None):
        """
        构造器
        :param filename: 文件名
        :param sheetname: sheet名
        """
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        """
        获取所有测试用例
        :return:所有用例
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_data_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        one_list = []
        for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_data_tuple, one_tuple)))
        return one_list

    def get_case(self, row):
        """
        获取单个测试用例
        :return:单条用例
        """
        return self.get_cases()[row - 1]

    def write_result(self, row, actual, result):
        """
        写入结果，注意同一个workbook对象，如果将数据写入到多个表单中，那么只有最后一个表单能写入成功
        :param row:行号
        :param actual:实际值
        :param result:结果
        :return:
        """
        other_wb = load_workbook(self.filename)
        actual_col = testcase_config.get_int("column", "actual")
        result_col = testcase_config.get_int("column", "result")
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=actual_col, value=actual)
            other_ws.cell(row=row, column=result_col, value=result)
            other_wb.save(self.filename)
        else:
            print("传入的行号有误，行号应为大于1的整数")


filename = os.path.join(DATA_DIR, testcase_config.get_value("file path", "cases_path"))
if __name__ == '__main__':
    do_excel = HandleExcel(filename, "add")
    cases = do_excel.get_cases()
    print(cases)
    case = do_excel.get_case(1)
    print(case)
    do_excel.write_result(2, 12, "Pass")
    pass
